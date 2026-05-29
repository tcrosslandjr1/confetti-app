"""
Excel Writer — NAS Storage
===========================
Writes/appends venue data to a master Excel workbook on the Synology NAS.
Workbook has one sheet per category + a "All Venues" master sheet.
Never overwrites existing rows — only appends new ones.
Tyrone can manually add venues to any sheet; they'll be picked up on next sync.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Run: pip install openpyxl")

log = logging.getLogger(__name__)

# Column definitions for each venue sheet
COLUMNS = [
    ("place_id",            "Place ID"),
    ("name",                "Name"),
    ("category",            "Category"),
    ("subcategory",         "Subcategory"),
    ("city",                "City"),
    ("neighborhood",        "Neighborhood"),
    ("address",             "Address"),
    ("price_range",         "Price"),
    ("google_rating",       "Google Rating"),
    ("google_review_count", "Google Reviews"),
    ("yelp_rating",         "Yelp Rating"),
    ("yelp_review_count",   "Yelp Reviews"),
    ("tiktok_mention_count","TikTok Mentions"),
    ("trending_score",      "Trending Score"),
    ("is_trending",         "Trending?"),
    ("is_featured",         "Featured?"),
    ("tags",                "Tags"),
    ("phone",               "Phone"),
    ("website",             "Website"),
    ("description",         "Description"),
    ("data_sources",        "Sources"),
    ("manually_added",      "Manual?"),
    ("curator_notes",       "Notes (Tyrone)"),
    ("last_fetched_at",     "Last Fetched"),
]

HEADER_FILL  = PatternFill("solid", fgColor="1C1C2E") if OPENPYXL_AVAILABLE else None
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11) if OPENPYXL_AVAILABLE else None
ACCENT_FILL  = PatternFill("solid", fgColor="FF6B6B") if OPENPYXL_AVAILABLE else None  # Confetti pink
ALT_ROW_FILL = PatternFill("solid", fgColor="F8F8FF") if OPENPYXL_AVAILABLE else None


def _get_or_create_workbook(path: str) -> "openpyxl.Workbook":
    """Load existing workbook or create a new one."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        log.info(f"Loaded existing workbook: {path}")
    else:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        log.info(f"Created new workbook: {path}")
    return wb


def _get_or_create_sheet(wb: "openpyxl.Workbook", sheet_name: str) -> "openpyxl.worksheet.worksheet.Worksheet":
    """Get existing sheet or create with headers."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    _write_header(ws)
    return ws


def _write_header(ws):
    """Write styled header row."""
    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # Set column widths
    widths = {
        1: 28,   # place_id
        2: 30,   # name
        3: 14,   # category
        4: 18,   # subcategory
        5: 12,   # city
        6: 18,   # neighborhood
        7: 35,   # address
        8: 7,    # price
        9: 13,   # google rating
        10: 15,  # google reviews
        11: 12,  # yelp rating
        12: 14,  # yelp reviews
        13: 16,  # tiktok
        14: 14,  # trending score
        15: 10,  # trending?
        16: 10,  # featured?
        17: 30,  # tags
        18: 16,  # phone
        19: 35,  # website
        20: 60,  # description
        21: 16,  # sources
        22: 10,  # manual?
        23: 40,  # notes
        24: 20,  # last fetched
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def _get_existing_place_ids(ws) -> set:
    """Return all place_ids already in the sheet (column 1)."""
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    return ids


def _get_existing_names(ws) -> set:
    """Return all venue names already in the sheet (column 2, lowercased)."""
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            names.add(str(row[1]).lower().strip())
    return names


def _format_cell_value(key: str, value) -> Any:
    """Convert value to Excel-friendly format."""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return value


def _append_venue_row(ws, venue: Dict[str, Any], row_num: int):
    """Append one venue as a row."""
    fill = ALT_ROW_FILL if row_num % 2 == 0 else None

    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        raw = venue.get(key)
        val = _format_cell_value(key, raw)
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = Alignment(wrap_text=False, vertical="center")
        if fill:
            cell.fill = fill

        # Highlight trending venues
        if key == "is_trending" and raw is True:
            cell.fill = PatternFill("solid", fgColor="FFD700")
            cell.font = Font(bold=True)

        # Color-code trending score
        if key == "trending_score" and raw:
            if raw >= 80:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif raw >= 60:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
            elif raw < 30:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")


def write_venues_to_excel(
    venues_by_category: Dict[str, List[Dict[str, Any]]],
    excel_path: str,
) -> int:
    """
    Write/append venues to the Excel file on NAS.
    - One sheet per category
    - "All Venues" master sheet
    - Never overwrites existing rows
    Returns total new rows written.
    """
    if not OPENPYXL_AVAILABLE:
        log.error("openpyxl not available — cannot write Excel")
        return 0

    wb = _get_or_create_workbook(excel_path)
    total_new = 0

    # Flatten all venues for the master sheet
    all_venues_flat = []

    for category, venues in venues_by_category.items():
        sheet_name = category.replace("_", " ").title()[:31]  # Excel limit: 31 chars
        ws = _get_or_create_sheet(wb, sheet_name)

        existing_ids   = _get_existing_place_ids(ws)
        existing_names = _get_existing_names(ws)
        current_row    = ws.max_row + 1

        new_in_cat = 0
        for venue in venues:
            pid  = str(venue.get("place_id", "")) if venue.get("place_id") else ""
            name = venue.get("name", "").lower().strip()

            # Skip if already exists
            if (pid and pid in existing_ids) or name in existing_names:
                continue

            _append_venue_row(ws, venue, current_row)
            existing_ids.add(pid)
            existing_names.add(name)
            all_venues_flat.append(venue)
            current_row += 1
            new_in_cat  += 1

        log.info(f"Excel [{sheet_name}]: {new_in_cat} new rows appended")
        total_new += new_in_cat

    # Update / create the master "All Venues" sheet
    master_name = "All Venues"
    ws_all = _get_or_create_sheet(wb, master_name)
    existing_all_ids   = _get_existing_place_ids(ws_all)
    existing_all_names = _get_existing_names(ws_all)
    master_row = ws_all.max_row + 1

    for venue in all_venues_flat:
        pid  = str(venue.get("place_id", "")) if venue.get("place_id") else ""
        name = venue.get("name", "").lower().strip()
        if (pid and pid in existing_all_ids) or name in existing_all_names:
            continue
        _append_venue_row(ws_all, venue, master_row)
        existing_all_ids.add(pid)
        existing_all_names.add(name)
        master_row += 1

    # Move "All Venues" to first position
    if master_name in wb.sheetnames:
        wb.move_sheet(master_name, offset=-(wb.sheetnames.index(master_name)))

    try:
        wb.save(excel_path)
        log.info(f"Excel saved: {excel_path} ({total_new} new rows total)")
    except Exception as e:
        log.error(f"Failed to save Excel: {e}")

    return total_new


def read_manual_additions(excel_path: str) -> List[Dict[str, Any]]:
    """
    Read any manually-added venues from the Excel file.
    Rows where 'Manual?' = 'Yes' are returned for Supabase sync.
    This lets Tyrone add places directly to the spreadsheet.
    """
    if not OPENPYXL_AVAILABLE or not os.path.exists(excel_path):
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    manual_venues = []
    col_keys = [k for k, _ in COLUMNS]

    # Check every sheet for manually-added rows
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            venue = {col_keys[i]: row[i] for i in range(min(len(col_keys), len(row)))}
            # manually_added column
            if str(venue.get("manually_added", "")).lower() in ("yes", "true", "1"):
                venue["manually_added"] = True
                venue["category"] = sheet_name.lower().replace(" ", "_")
                manual_venues.append(venue)

    log.info(f"Excel: {len(manual_venues)} manually-added venues found")
    return manual_venues
