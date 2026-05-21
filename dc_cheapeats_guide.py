from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styling ──
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2D2D2D')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Arial', size=10)
data_align = Alignment(vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
mid_header_fill = PatternFill('solid', fgColor='1B4332')
price_tag_font = Font(name='Arial', bold=True, size=10, color='1B4332')

# Category fills
cat_fills = {
    "Caribbean/Jerk": PatternFill('solid', fgColor='FFF3E0'),
    "Seafood": PatternFill('solid', fgColor='E3F2FD'),
    "Tacos/Mexican": PatternFill('solid', fgColor='FFF9C4'),
    "Fried Chicken": PatternFill('solid', fgColor='FFEBEE'),
    "Pizza/Italian": PatternFill('solid', fgColor='F3E5F5'),
    "Indian": PatternFill('solid', fgColor='E8F5E9'),
    "African/Suya": PatternFill('solid', fgColor='FCE4EC'),
    "Wings": PatternFill('solid', fgColor='FBE9E7'),
    "Hibachi/Asian": PatternFill('solid', fgColor='E0F7FA'),
    "Burgers": PatternFill('solid', fgColor='F1F8E9'),
    "Sandwiches/Deli": PatternFill('solid', fgColor='FFFDE7'),
    "Breakfast/Brunch": PatternFill('solid', fgColor='FFF8E1'),
    "Chinese": PatternFill('solid', fgColor='E8EAF6'),
    "Bakery/Desserts": PatternFill('solid', fgColor='FCE4EC'),
    "Soul Food/Southern": PatternFill('solid', fgColor='EFEBE9'),
    "Halal": PatternFill('solid', fgColor='E0F2F1'),
    "BBQ": PatternFill('solid', fgColor='FBE9E7'),
    "Persian/Middle Eastern": PatternFill('solid', fgColor='F3E5F5'),
    "French/European": PatternFill('solid', fgColor='EDE7F6'),
    "New American": PatternFill('solid', fgColor='E3F2FD'),
    "Mediterranean": PatternFill('solid', fgColor='E8F5E9'),
    "Ethiopian/East African": PatternFill('solid', fgColor='FFF3E0'),
    "West African": PatternFill('solid', fgColor='FCE4EC'),
    "Thai": PatternFill('solid', fgColor='E0F7FA'),
    "Japanese/Sushi": PatternFill('solid', fgColor='E8EAF6'),
    "Korean": PatternFill('solid', fgColor='F1F8E9'),
    "Vietnamese": PatternFill('solid', fgColor='E0F2F1'),
    "Latin/Brazilian": PatternFill('solid', fgColor='FFF9C4'),
    "Afghan": PatternFill('solid', fgColor='EFEBE9'),
    "Ukrainian": PatternFill('solid', fgColor='E3F2FD'),
    "Gastropub": PatternFill('solid', fgColor='FBE9E7'),
    "Indian-Nepalese": PatternFill('solid', fgColor='E8F5E9'),
    "Salvadoran/Fusion": PatternFill('solid', fgColor='FFF3E0'),
}

# ── DATA ──
# CHEAP EATS (from @dccheapeats TikTok): [Name, Cuisine, Known For, Location, Area, Views, Notes]
cheap_eats = [
    ["(Jerk Spot - Pinned)", "Caribbean/Jerk", "Best Jerk in Maryland", "Maryland", "MD", "433.8K", "PINNED - Top pick, massive views"],
    ["The Seafood Trap (Pinned)", "Seafood", "Big seafood platters", "DMV Area", "DC/MD/VA", "487.4K", "PINNED - Big-portion seafood"],
    ["(Taco Spot - Pinned)", "Tacos/Mexican", "Best tacos in the DMV", "DMV Area", "DC/MD/VA", "578.8K", "PINNED - Most viewed, fan favorite"],
    ["The Block Oven", "Pizza/Italian", "New pizza oven spot", "Northern Virginia", "VA", "1K", "Just opened — get in early"],
    ["(Pikesville Spot)", "Various", "Worth the drive", "Pikesville, MD", "MD", "2.4K", "Drove from DC just to try it"],
    ["(New Fried Chicken)", "Fried Chicken", "New fried chicken spot", "Washington, DC", "DC", "6.2K", "New opening in DC"],
    ["(Indian Hidden Gem)", "Indian", "Hidden gem Indian food", "Washington, DC", "DC", "2.4K", "Under-the-radar Indian spot"],
    ["Catalano", "Pizza/Italian", "Crab pizza & crab rolls", "DMV Area", "DC/MD/VA", "26.9K", "Unique crab pizza combo"],
    ["MAD Seafood", "Halal", "New halal seafood", "Northern Virginia", "VA", "11.4K", "Halal seafood — rare find in NoVA"],
    ["Joe's (Wings Pt 14)", "Wings", "Best wings in DC series", "Washington, DC", "DC", "9.6K", "Part of ongoing best-wings series"],
    ["Prime Grill", "African/Suya", "Best suya wrap in DMV", "DMV Area", "DC/MD/VA", "2K", "Suya wrap standout"],
    ["Flame Japanese Hibachi", "Hibachi/Asian", "Takeout hibachi $9-$15", "Laurel, MD", "MD", "54.5K", "Affordable hibachi, takeout only"],
    ["Al's Famous Delicatessen", "Sandwiches/Deli", "Bulgogi cheesesteak", "Washington, DC", "DC", "11.3K", "Korean-American fusion deli"],
    ["(Best Chicken DC)", "Fried Chicken", "Best chicken in DC", "Washington, DC", "DC", "20.3K", "New spot, strong reviews"],
    ["(Massive Sandwich Spot)", "Sandwiches/Deli", "Massive loaded sandwiches", "DMV Area", "DC/MD/VA", "16.8K", "Huge portions, went viral"],
    ["(Beltsville Suya Gem)", "African/Suya", "$10 suya plates", "Beltsville, MD", "MD", "205.8K", "Hidden gem, massive views — $10 suya"],
    ["The Jerk Pit", "Caribbean/Jerk", "Oxtail $11.95, curry goat, chicken $9", "DMV Area", "DC/MD/VA", "100.3K", "Incredible value — oxtail under $12"],
    ["Sunrise Caribbean Restaurant", "Caribbean/Jerk", "Caribbean food, new Greenbelt location", "Greenbelt, MD", "MD", "13.2K", "Just moved/expanded to Greenbelt"],
    ["Becky's Seafood House", "Seafood", "New seafood takeout spot", "Washington, DC", "DC", "79.2K", "Takeout seafood, very popular"],
    ["Keith and Sons", "Soul Food/Southern", "Classic soul food", "DMV Area", "DC/MD/VA", "28.8K", "If you know you know — local favorite"],
    ["Crimson Coward", "Fried Chicken", "Nashville hot chicken chain", "DMV Area", "DC/MD/VA", "4.7K", "Nashville hot chicken comes to DMV"],
    ["Roaming Rooster x Capital City Mambo", "Fried Chicken", "Collab — chicken + mambo sauce", "Washington, DC", "DC", "3.5K", "DC icons collab — must try"],
    ["Smokey's", "BBQ", "Favorite spot in DC", "Washington, DC", "DC", "16.3K", "Creator's personal favorite"],
    ["China Boy", "Chinese", "Best Chinese in Chinatown DC", "Chinatown, DC", "DC", "6.7K", "Chinatown staple"],
    ["Langston Golf Course", "Breakfast/Brunch", "Best pancakes in DC", "Washington, DC", "DC", "57.7K", "Unexpected pancake gem at a golf course"],
    ["Rita's Italian Ice", "Bakery/Desserts", "Free Italian ice promos", "DMV Area", "DC/MD/VA", "8.7K", "Watch for free giveaway days"],
    ["GW Delicatessen", "Sandwiches/Deli", "Best bagel sandwich in DC", "Washington, DC", "DC", "11.2K", "Bagel sandwiches, classic deli"],
    ["Alandin", "Bakery/Desserts", "Best cookies in the DMV", "DMV Area", "DC/MD/VA", "2.8K", "Cookie destination"],
    ["Jon's Joint", "Burgers", "New burger joint", "Alexandria, VA", "VA", "190.2K", "Massive views — burger spot in Old Town area"],
    ["Buffalo Wild Wings Go", "Wings", "Fast-casual wings", "Maryland", "MD", "4.7K", "New BWW Go format in MD"],
    ["(Indian Hidden Gem #2)", "Indian", "Hidden gem Indian food", "Washington, DC", "DC", "5.6K", "Another under-the-radar Indian find"],
    ["IHOP", "Breakfast/Brunch", "Free pancake promos", "DMV Area", "DC/MD/VA", "766.7K", "Free pancake day — most viral video"],
    ["(Authentic Taco Spot)", "Tacos/Mexican", "Most authentic tacos in MD", "Maryland", "MD", "75.3K", "Legit Mexican tacos in Maryland"],
    ["Brookland Grill", "Soul Food/Southern", "Fried fish", "Brookland, DC", "DC", "25.3K", "Neighborhood fried fish spot"],
    ["Z Burger & Maman Joon", "Persian/Middle Eastern", "Burgers + Persian kabob", "1400 St NW, DC", "DC", "22.8K", "Two-in-one: burgers & Persian food"],
]

# MID-LEVEL RESTAURANTS: [Name, Cuisine, Known For, Location, Area, Price Range, Notes]
mid_level = [
    # DC — Date Night / Social Dining
    ["Le Diplomate", "French/European", "French bistro, steak frites, brunch scene", "14th St NW, DC", "DC", "$$$", "DC's most iconic French bistro — always packed"],
    ["St. Anselm", "New American", "Wood-grilled meats, rib steak, craft cocktails", "Union Market, DC", "DC", "$$$", "One of DC's best steakhouse values"],
    ["Old Ebbitt Grill", "New American", "Oysters, burgers, historic bar — est. 1856", "15th St NW, DC", "DC", "$$$", "DC landmark near the White House"],
    ["Founding Farmers DC", "New American", "Farm-to-table, fried chicken, brunch", "Pennsylvania Ave NW, DC", "DC", "$$", "Consistently packed — weekend brunch is legendary"],
    ["The Hamilton", "New American", "Live music venue + full restaurant", "14th St NW, DC", "DC", "$$", "Dinner and a show under one roof"],
    ["L'Ardente", "Pizza/Italian", "Handmade pasta, Italian fine-casual", "City Center, DC", "DC", "$$$", "Upscale Italian with incredible pastas"],
    ["Unconventional Diner", "New American", "Elevated comfort food, mac & cheese, brunch", "Shaw, DC", "DC", "$$", "Creative takes on diner classics"],
    ["Lapis", "Afghan", "Kofta, bolani, Afghan kabobs", "Adams Morgan, DC", "DC", "$$", "Best Afghan food in the city — cozy vibe"],
    ["Zaytinya", "Mediterranean", "Mezze, Turkish-Greek-Lebanese small plates", "Gallery Place, DC", "DC", "$$$", "José Andrés — world-class mezze"],
    ["Dear Sushi", "Japanese/Sushi", "Omakase under $50", "DC", "DC", "$$", "Rare affordable omakase experience"],
    ["Tapori", "Indian-Nepalese", "Chaat, kebabs, Nepalese flavors", "H Street, DC", "DC", "$$", "H Street hidden gem — incredible chaat"],
    ["Providencia", "Salvadoran/Fusion", "Salvadoran-Dominican-Taiwanese fusion", "H Street, DC", "DC", "$$", "Tiny spot, massive flavors — totally unique"],
    ["Minetta Tavern", "French/European", "Classic French-American, steak, cocktails", "Penn Quarter, DC", "DC", "$$$", "NYC import — sophisticated date night"],

    # DC — Black-Owned / Cultural
    ["Dōgon", "West African", "Seychellois flavors, crab cakes w/ coconut, lamb lollipops", "Shaw, DC", "DC", "$$$", "Stunning rooftop — Black-owned fine dining"],
    ["Vortex Restaurant & Lounge", "Ethiopian/East African", "East/West African & Caribbean-inspired, 3 levels", "L St NW, DC", "DC", "$$", "Black & woman-owned — great vibe"],
    ["Hedzole", "West African", "Jollof rice bowls, West African comfort", "DC", "DC", "$$", "Fast-casual West African — chef chats w/ every diner"],
    ["Bukom Café", "West African", "West African classics, live reggae weekends", "Adams Morgan, DC", "DC", "$$", "Live music + outstanding food — neighborhood staple"],
    ["Milk & Honey", "Breakfast/Brunch", "NOLA-style brunch, shrimp & grits, Biscoff waffles", "The Wharf, DC", "DC", "$$", "Must-visit brunch — New Orleans themed"],
    ["Chercher Ethiopian", "Ethiopian/East African", "Ethiopian platters, tibs, kitfo", "DC", "DC", "$$", "Top Ethiopian in the city — affordable and fresh"],
    ["A&J Restaurant", "Chinese", "Northern Chinese dim sum, dumplings, dan-dan noodles", "Annandale/Rockville", "DC/MD/VA", "$", "Old-school dim sum standby — multiple locations"],

    # Maryland
    ["Aventino", "Pizza/Italian", "Roman-inspired, antipasti, handmade pasta", "Bethesda, MD", "MD", "$$$", "Chef Mike Friedman — Eater 38 Best 2026"],
    ["Ruta", "Ukrainian", "Varenyky, chicken Kyiv, holubtsi", "Bethesda, MD", "MD", "$$", "Rare Ukrainian restaurant — Eater 38 Best 2026"],
    ["J. Hollinger's Waterman's Chophouse", "Seafood", "Chef-driven seafood, epic Sunday brunch buffet", "Silver Spring, MD", "MD", "$$$", "Brunch: ribeye, hot honey chicken, crab deviled eggs"],
    ["All Set", "New American", "Lobster rolls, MD crab cakes, seasonal menu", "Silver Spring, MD", "MD", "$$", "Black-owned — modern seasonal dining"],
    ["El Viejo", "Tacos/Mexican", "Pupusas, Salvadoran-Mexican street food", "Silver Spring, MD", "MD", "$", "Washingtonian 100 Best — incredible pupusas"],
    ["2Fifty Texas BBQ", "BBQ", "Central Texas brisket, ribs, scratch sides", "Riverdale, MD", "MD", "$$", "Nationally acclaimed — long lines worth it"],
    ["Cielo Rojo", "Tacos/Mexican", "Cali-inspired Mexican, tacos, enchiladas", "Takoma Park, MD", "MD", "$$", "Eater 38 Best 2026 — great small plates"],

    # Virginia
    ["Tio Pelé", "Latin/Brazilian", "Brazilian-Latin sharing plates, grilled meats, cocktails", "Arlington, VA", "VA", "$$", "Bold flavors — new in Arlington"],
    ["Dok Khao", "Thai", "Panang curry, pad see ew, papaya salads", "Alexandria, VA", "VA", "$$", "Authentic Thai in Old Town — standout curries"],
    ["Chao Ban", "Vietnamese", "Pho, rice platters, Vietnamese-American fusion", "Tysons, VA", "VA", "$$", "Gulf South + Mid-Atlantic Vietnamese twist"],
    ["Peter Chang", "Chinese", "Legendary Szechuan, dry-fried eggplant, dan-dan", "Multiple VA locations", "VA", "$$", "Chef Peter Chang — best Szechuan in the DMV"],
    ["Grazie Nonna", "Pizza/Italian", "Italian-American comfort, pasta, chicken parm", "Old Town Alexandria, VA", "VA", "$$", "Coming 2026 — Italian-American done right"],
    ["Maman", "French/European", "French café, pastries, seasonal plates", "Crystal City, VA", "VA", "$$", "NYC French café chain — great for brunch/coffee"],

    # Date Night / Special Occasion
    ["The Fountain Inn", "New American", "New arrival — seasonal tasting vibes", "DC", "DC", "$$$", "Resy hit list — newest date night spot"],
    ["Chloe", "New American", "Modern American small plates, wine bar", "Navy Yard, DC", "DC", "$$$", "Intimate wine bar / small plates — Navy Yard gem"],
    ["Amelie", "French/European", "French wine bar, charcuterie, cocktails", "DC", "DC", "$$$", "Parisian wine bar vibes"],
    ["CODE RED", "Japanese/Sushi", "Japanese-inspired, creative cocktails", "DC", "DC", "$$$", "Trendy — great for group outings"],
    ["Astoria DC", "Mediterranean", "Mediterranean small plates, hookah vibes", "DC", "DC", "$$", "Social dining — hookah + mezze plates"],
    ["Albi", "Mediterranean", "Wood-fired Levantine, lamb, hummus", "Navy Yard, DC", "DC", "$$$", "Washingtonian 100 Best — fire-cooked everything"],
    ["Yellow Cuisine", "Persian/Middle Eastern", "Persian-inspired, Michelin Guide pick", "DC", "DC", "$$", "Michelin Guide 2025 value pick"],
    ["Your Only Friend", "Gastropub", "Michelin-recommended gastropub, craft cocktails", "Petworth, DC", "DC", "$$", "Michelin Guide 2025 — neighborhood gem"],
]


# ═══════════════════════════════════════════════════════════
# SHEET 1: ALL SPOTS (Cheap Eats)
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Cheap Eats"

headers = ["#", "Restaurant Name", "Cuisine / Category", "Known For", "Location", "Area", "Views", "Vibe / Notes"]
col_widths = [5, 28, 22, 35, 20, 18, 12, 35]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

for i, row in enumerate(cheap_eats, 1):
    ws.cell(row=i+1, column=1, value=i).font = data_font
    ws.cell(row=i+1, column=1).alignment = Alignment(horizontal='center')
    for j, val in enumerate(row, 2):
        cell = ws.cell(row=i+1, column=j, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        cuisine = row[1]
        if cuisine in cat_fills:
            ws.cell(row=i+1, column=3).fill = cat_fills[cuisine]

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:H{len(cheap_eats)+1}"

# ═══════════════════════════════════════════════════════════
# SHEET 2: MID-LEVEL RESTAURANTS
# ═══════════════════════════════════════════════════════════
ws_mid = wb.create_sheet("Mid-Level")

mid_headers = ["#", "Restaurant Name", "Cuisine / Category", "Known For", "Location", "Area", "Price", "Vibe / Notes"]
mid_widths = [5, 30, 24, 40, 24, 10, 8, 42]

for col, (header, width) in enumerate(zip(mid_headers, mid_widths), 1):
    cell = ws_mid.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = mid_header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_mid.column_dimensions[get_column_letter(col)].width = width

for i, row in enumerate(mid_level, 1):
    ws_mid.cell(row=i+1, column=1, value=i).font = data_font
    ws_mid.cell(row=i+1, column=1).alignment = Alignment(horizontal='center')
    for j, val in enumerate(row, 2):
        cell = ws_mid.cell(row=i+1, column=j, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        cuisine = row[1]
        if cuisine in cat_fills:
            ws_mid.cell(row=i+1, column=3).fill = cat_fills[cuisine]
    # Bold the price column
    ws_mid.cell(row=i+1, column=7).font = price_tag_font

ws_mid.freeze_panes = 'A2'
ws_mid.auto_filter.ref = f"A1:H{len(mid_level)+1}"

# ═══════════════════════════════════════════════════════════
# SHEET 3: FULL MASTER LIST (all spots combined)
# ═══════════════════════════════════════════════════════════
ws_all = wb.create_sheet("All Spots Combined")

all_headers = ["#", "Restaurant Name", "Cuisine / Category", "Known For", "Location", "Area", "Tier", "Notes"]
all_widths = [5, 30, 24, 40, 24, 10, 14, 42]

for col, (header, width) in enumerate(zip(all_headers, all_widths), 1):
    cell = ws_all.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill('solid', fgColor='1A1A2E')
    cell.alignment = header_align
    cell.border = thin_border
    ws_all.column_dimensions[get_column_letter(col)].width = width

cheap_fill = PatternFill('solid', fgColor='FFF8E1')
mid_fill_bg = PatternFill('solid', fgColor='E8F5E9')

idx = 1
for row in cheap_eats:
    r = idx + 1
    ws_all.cell(row=r, column=1, value=idx).font = data_font
    ws_all.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws_all.cell(row=r, column=2, value=row[0]).font = data_font
    ws_all.cell(row=r, column=3, value=row[1]).font = data_font
    ws_all.cell(row=r, column=4, value=row[2]).font = data_font
    ws_all.cell(row=r, column=5, value=row[3]).font = data_font
    ws_all.cell(row=r, column=6, value=row[4]).font = data_font
    ws_all.cell(row=r, column=7, value="$ Cheap Eats").font = Font(name='Arial', size=10, color='E65100')
    ws_all.cell(row=r, column=7).fill = cheap_fill
    ws_all.cell(row=r, column=8, value=row[6]).font = data_font
    for c in range(1, 9):
        ws_all.cell(row=r, column=c).border = thin_border
        ws_all.cell(row=r, column=c).alignment = data_align
    idx += 1

for row in mid_level:
    r = idx + 1
    ws_all.cell(row=r, column=1, value=idx).font = data_font
    ws_all.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws_all.cell(row=r, column=2, value=row[0]).font = data_font
    ws_all.cell(row=r, column=3, value=row[1]).font = data_font
    ws_all.cell(row=r, column=4, value=row[2]).font = data_font
    ws_all.cell(row=r, column=5, value=row[3]).font = data_font
    ws_all.cell(row=r, column=6, value=row[4]).font = data_font
    ws_all.cell(row=r, column=7, value=f"{row[5]} Mid-Level").font = Font(name='Arial', size=10, color='1B4332')
    ws_all.cell(row=r, column=7).fill = mid_fill_bg
    ws_all.cell(row=r, column=8, value=row[6]).font = data_font
    for c in range(1, 9):
        ws_all.cell(row=r, column=c).border = thin_border
        ws_all.cell(row=r, column=c).alignment = data_align
    idx += 1

ws_all.freeze_panes = 'A2'
ws_all.auto_filter.ref = f"A1:H{idx}"

# ═══════════════════════════════════════════════════════════
# SHEET 4: BY AREA
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("By Area")
all_restaurants = [(r, "Cheap Eats") for r in cheap_eats] + [(r, "Mid-Level") for r in mid_level]

area_groups = {"DC": [], "MD": [], "VA": [], "DC/MD/VA": []}
for r, tier in all_restaurants:
    area = r[4]
    if area in area_groups:
        area_groups[area].append((r, tier))

row_idx = 1
for area, spots in area_groups.items():
    if not spots:
        continue
    cell = ws3.cell(row=row_idx, column=1, value=f"── {area} ({len(spots)} spots) ──")
    cell.font = Font(name='Arial', bold=True, size=12, color='2D2D2D')
    row_idx += 1
    for h_col, h_text in enumerate(["Restaurant", "Cuisine", "Known For", "Tier"], 1):
        c = ws3.cell(row=row_idx, column=h_col, value=h_text)
        c.font = header_font
        c.fill = header_fill
    row_idx += 1
    for spot, tier in spots:
        ws3.cell(row=row_idx, column=1, value=spot[0]).font = data_font
        ws3.cell(row=row_idx, column=2, value=spot[1]).font = data_font
        ws3.cell(row=row_idx, column=3, value=spot[2]).font = data_font
        ws3.cell(row=row_idx, column=4, value=tier).font = data_font
        if tier == "Mid-Level":
            ws3.cell(row=row_idx, column=4).fill = mid_fill_bg
        else:
            ws3.cell(row=row_idx, column=4).fill = cheap_fill
        row_idx += 1
    row_idx += 1

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 24
ws3.column_dimensions['C'].width = 40
ws3.column_dimensions['D'].width = 14

# ═══════════════════════════════════════════════════════════
# SHEET 5: DATE NIGHT PICKS
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Date Night Picks")
ws5.cell(row=1, column=1, value="DMV DATE NIGHT PICKS — CHEAP TO MID-LEVEL").font = Font(name='Arial', bold=True, size=14, color='880E4F')
ws5.merge_cells('A1:F1')

date_night_spots = [r for r in mid_level if any(kw in r[6].lower() for kw in ['date', 'romantic', 'wine', 'intimate', 'rooftop', 'bistro', 'cocktail', 'vibe', 'social', 'trendy', 'hookah', 'parisian', 'sophisticated'])]

for h_col, h_text in enumerate(["#", "Restaurant", "Cuisine", "Known For", "Price", "Why It Works"], 1):
    c = ws5.cell(row=3, column=h_col, value=h_text)
    c.font = header_font
    c.fill = PatternFill('solid', fgColor='880E4F')
    c.alignment = header_align

for i, spot in enumerate(date_night_spots, 1):
    row = i + 3
    ws5.cell(row=row, column=1, value=i).font = Font(name='Arial', bold=True, size=11)
    ws5.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws5.cell(row=row, column=2, value=spot[0]).font = data_font
    ws5.cell(row=row, column=3, value=spot[1]).font = data_font
    ws5.cell(row=row, column=4, value=spot[2]).font = data_font
    ws5.cell(row=row, column=5, value=spot[5]).font = price_tag_font
    ws5.cell(row=row, column=6, value=spot[6]).font = data_font

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 22
ws5.column_dimensions['D'].width = 40
ws5.column_dimensions['E'].width = 8
ws5.column_dimensions['F'].width = 45

outpath = '/sessions/cool-hopeful-gauss/mnt/ai-lifestyle-concierge/DC_CheapEats_Food_Guide.xlsx'
wb.save(outpath)
print(f"Saved to {outpath}")
print(f"Cheap eats: {len(cheap_eats)}")
print(f"Mid-level:  {len(mid_level)}")
print(f"TOTAL:      {len(cheap_eats) + len(mid_level)}")
