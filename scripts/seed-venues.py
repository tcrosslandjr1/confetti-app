#!/usr/bin/env python3
"""
🎊 Confetti Venue Seeder
========================
Reads all Excel guide files, converts to structured venue data, and outputs:
  1. A Supabase SQL migration (creates `venues` table + inserts all records)
  2. A TypeScript venue-knowledge module the AI agents import directly

Run:
  python3 scripts/seed-venues.py

Outputs:
  supabase/migrations/20260524000000_seed_venue_knowledge.sql
  src/lib/agents/venue-knowledge.ts
"""

import os, json, re, glob, hashlib
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
GUIDE_PATTERN = str(ROOT / "*_Guide*.xlsx")

# ─── City geo coordinates for proximity search ─────────────────
CITY_COORDS = {
    "DC": (38.9072, -77.0369),
    "DC CheapEats Food": (38.9072, -77.0369),
    "Maryland": (39.0458, -76.6413),
    "Virginia": (38.8462, -77.3064),
    "NYC": (40.7128, -74.0060),
    "NYC SexyDatesNYC": (40.7128, -74.0060),
    "NewJersey": (40.0583, -74.4057),
    "Miami": (25.7617, -80.1918),
    "FortLauderdale": (26.1224, -80.1373),
    "Chicago": (41.8781, -87.6298),
    "Atlanta": (33.7490, -84.3880),
    "Nashville": (36.1627, -86.7816),
    "Boston": (42.3601, -71.0589),
    "Seattle": (47.6062, -122.3321),
    "Philadelphia": (39.9526, -75.1652),
    "Austin": (30.2672, -97.7431),
    "Denver": (39.7392, -104.9903),
    "NewOrleans": (29.9511, -90.0715),
    "Toronto": (43.6532, -79.3832),
    "SF": (37.7749, -122.4194),
    "California": (34.0522, -118.2437),
    "Cincinnati DateNight": (39.1031, -84.5120),
    "LA LaLaGuide": (34.0522, -118.2437),
}

# ─── Category → Occasion mapping ──────────────────────────────
CATEGORY_TO_OCCASION = {
    "Date Night": "date-night",
    "Group Night Out": "group-night-out",
    "In-Laws / Family": "family",
    "In-Laws - Family": "family",
    "Girls Night Out": "girls-night",
    "Girls Night": "girls-night",
    "Guys Night": "guys-night",
    "Guys Night Out": "guys-night",
    "Bachelor Party": "bachelor",
    "Bachelorette": "bachelorette",
    "Squad Outside": "outdoor-adventure",
    "Outdoor Activities": "outdoor-adventure",
    "Brunch": "brunch",
    "Late Night": "late-night",
    "Cheap Eats": "budget-friendly",
    "Fine Dining": "special-occasion",
    "Rooftops": "rooftop",
    "Live Music": "live-music",
    "Cocktail Bars": "cocktails",
    "Sports Bars": "sports",
}


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def price_to_level(price: str) -> int:
    if not price:
        return 2
    count = price.count("$")
    return min(max(count, 1), 4)


def extract_city_from_filename(filename: str) -> str:
    base = Path(filename).stem
    for suffix in ["_Confetti_Guide", "_Guide", "_Spots", "_v2"]:
        base = base.replace(suffix, "")
    return base.replace("_", " ").strip()


def normalize_city(city: str) -> str:
    """Normalize to clean city name for DB"""
    mapping = {
        "DC": "Washington",
        "DC CheapEats Food": "Washington",
        "NYC": "New York",
        "NYC SexyDatesNYC": "New York",
        "SF": "San Francisco",
        "NewOrleans": "New Orleans",
        "NewJersey": "New Jersey",
        "FortLauderdale": "Fort Lauderdale",
        "Cincinnati DateNight": "Cincinnati",
        "LA LaLaGuide": "Los Angeles",
    }
    return mapping.get(city, city)


def stable_id(name: str, city: str, address: str) -> str:
    """Generate a stable UUID-like ID from venue identity"""
    raw = f"{name.lower().strip()}|{city.lower().strip()}|{address.lower().strip()}"
    h = hashlib.md5(raw.encode()).hexdigest()
    return f"{h[:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"


def extract_vibe_tags(vibe_notes: str, cuisine: str) -> list[str]:
    """Pull vibe/mood keywords from notes"""
    tags = []
    keywords = {
        "romantic": "romantic", "cozy": "cozy", "upscale": "upscale",
        "trendy": "trendy", "chill": "chill", "loud": "energetic",
        "intimate": "intimate", "elegant": "elegant", "fun": "fun",
        "viral": "viral", "hidden": "hidden-gem", "rooftop": "rooftop",
        "waterfront": "waterfront", "outdoor": "outdoor", "patio": "patio",
        "live music": "live-music", "dj": "dj", "dance": "dance",
        "cocktail": "cocktails", "wine": "wine", "whiskey": "whiskey",
        "brunch": "brunch", "late": "late-night", "sunset": "sunset",
        "instagram": "instagrammable", "tiktok": "viral",
        "michelin": "michelin", "iconic": "iconic",
    }
    combined = (vibe_notes + " " + cuisine).lower()
    for kw, tag in keywords.items():
        if kw in combined and tag not in tags:
            tags.append(tag)
    return tags


def parse_workbook(filepath: str) -> list[dict]:
    venues = []
    city_raw = extract_city_from_filename(filepath)
    city = normalize_city(city_raw)
    coords = CITY_COORDS.get(city_raw, (0, 0))

    wb = openpyxl.load_workbook(filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in ("overview", "summary", "notes", "sources"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if "name" in h:
                col_map["name"] = i
            elif "neighbor" in h:
                col_map["neighborhood"] = i
            elif "address" in h:
                col_map["address"] = i
            elif "type" in h or "cuisine" in h:
                col_map["cuisine"] = i
            elif "price" in h:
                col_map["price"] = i
            elif "vibe" in h or "note" in h:
                col_map["vibe"] = i
            elif "source" in h:
                col_map["source"] = i

        if "name" not in col_map:
            continue

        occasion = CATEGORY_TO_OCCASION.get(sheet_name.strip(), slugify(sheet_name))

        for row in rows[1:]:
            name = row[col_map["name"]] if col_map.get("name") is not None and row[col_map["name"]] else None
            if not name:
                continue

            name = str(name).strip()
            neighborhood = str(row[col_map.get("neighborhood", 0)] or "").strip()
            address = str(row[col_map.get("address", 0)] or "").strip()
            cuisine = str(row[col_map.get("cuisine", 0)] or "").strip()
            price = str(row[col_map.get("price", 0)] or "$$").strip()
            vibe = str(row[col_map.get("vibe", 0)] or "").strip()
            source = str(row[col_map.get("source", 0)] or "").strip()

            cuisine_tags = [t.strip() for t in cuisine.split("/") if t.strip()]
            vibe_tags = extract_vibe_tags(vibe, cuisine)

            venue = {
                "id": stable_id(name, city, address),
                "name": name,
                "slug": slugify(name),
                "city": city,
                "state": get_state(city_raw),
                "neighborhood": neighborhood,
                "address": address,
                "lat": coords[0],
                "lng": coords[1],
                "cuisine": cuisine,
                "cuisine_tags": cuisine_tags,
                "vibe_tags": vibe_tags,
                "occasion_tags": [occasion],
                "price": price,
                "price_level": price_to_level(price),
                "vibe_notes": vibe,
                "source_credit": source,
                "category": sheet_name.strip(),
                "sheet_occasion": occasion,
            }
            venues.append(venue)

    wb.close()
    return venues


def get_state(city_raw: str) -> str:
    state_map = {
        "DC": "DC", "DC CheapEats Food": "DC", "Maryland": "MD", "Virginia": "VA",
        "NYC": "NY", "NYC SexyDatesNYC": "NY", "NewJersey": "NJ",
        "Miami": "FL", "FortLauderdale": "FL",
        "Chicago": "IL", "Atlanta": "GA", "Nashville": "TN",
        "Boston": "MA", "Seattle": "WA", "Philadelphia": "PA",
        "Austin": "TX", "Denver": "CO", "NewOrleans": "LA",
        "Toronto": "ON", "SF": "CA", "California": "CA",
        "Cincinnati DateNight": "OH", "LA LaLaGuide": "CA",
    }
    return state_map.get(city_raw, "")


def deduplicate(venues: list[dict]) -> list[dict]:
    """Deduplicate, merging occasion tags for same venue"""
    seen = {}
    for v in venues:
        key = (v["name"].lower(), v["city"].lower())
        if key in seen:
            # Merge occasion tags
            existing = seen[key]
            for tag in v["occasion_tags"]:
                if tag not in existing["occasion_tags"]:
                    existing["occasion_tags"].append(tag)
            # Merge vibe tags
            for tag in v["vibe_tags"]:
                if tag not in existing["vibe_tags"]:
                    existing["vibe_tags"].append(tag)
        else:
            seen[key] = v
    return list(seen.values())


def escape_sql(s: str) -> str:
    return s.replace("'", "''")


def generate_migration(venues: list[dict]) -> str:
    """Generate Supabase SQL migration"""
    lines = [
        "-- ============================================================",
        "-- Confetti Venue Knowledge Base",
        f"-- Generated: {datetime.now().isoformat()}",
        f"-- Total venues: {len(venues)}",
        "-- ============================================================",
        "",
        "-- Create venues table (the AI agents' local knowledge)",
        "create table if not exists venues (",
        "  id            text primary key,",
        "  name          text not null,",
        "  slug          text not null,",
        "  city          text not null,",
        "  state         text,",
        "  neighborhood  text,",
        "  address       text,",
        "  lat           double precision,",
        "  lng           double precision,",
        "  cuisine       text,",
        "  cuisine_tags  text[] default '{}',",
        "  vibe_tags     text[] default '{}',",
        "  occasion_tags text[] default '{}',",
        "  price         text,",
        "  price_level   int default 2,",
        "  vibe_notes    text,",
        "  source_credit text,",
        "  rating        double precision,",
        "  rating_count  int default 0,",
        "  is_verified   boolean default false,",
        "  photo_url     text,",
        "  website       text,",
        "  phone         text,",
        "  hours         jsonb,",
        "  -- Learning fields",
        "  times_recommended int default 0,",
        "  times_accepted    int default 0,",
        "  times_rejected    int default 0,",
        "  avg_user_rating   double precision,",
        "  last_recommended  timestamptz,",
        "  popularity_score  double precision default 0,",
        "  created_at    timestamptz not null default now(),",
        "  updated_at    timestamptz not null default now()",
        ");",
        "",
        "-- Indexes for fast agent queries",
        "create index if not exists idx_venues_city on venues(city);",
        "create index if not exists idx_venues_city_occasion on venues using gin(occasion_tags);",
        "create index if not exists idx_venues_city_vibe on venues using gin(vibe_tags);",
        "create index if not exists idx_venues_city_cuisine on venues using gin(cuisine_tags);",
        "create index if not exists idx_venues_price on venues(price_level);",
        "create index if not exists idx_venues_popularity on venues(popularity_score desc);",
        "create index if not exists idx_venues_slug on venues(slug);",
        "",
        "-- Venue feedback table (how the system learns)",
        "create table if not exists venue_feedback (",
        "  id            uuid primary key default gen_random_uuid(),",
        "  venue_id      text not null references venues(id),",
        "  user_id       uuid not null,",
        "  action        text not null, -- 'accepted', 'rejected', 'saved', 'visited', 'rated'",
        "  rating        int, -- 1-5 if rated",
        "  context       jsonb default '{}', -- occasion, group_size, time_of_day, etc.",
        "  created_at    timestamptz not null default now()",
        ");",
        "",
        "create index if not exists idx_venue_feedback_venue on venue_feedback(venue_id);",
        "create index if not exists idx_venue_feedback_user on venue_feedback(user_id);",
        "",
        "-- Function to update venue popularity after feedback",
        "create or replace function update_venue_popularity()",
        "  returns trigger as $$",
        "begin",
        "  update venues set",
        "    times_recommended = times_recommended + (case when NEW.action = 'recommended' then 1 else 0 end),",
        "    times_accepted = times_accepted + (case when NEW.action = 'accepted' then 1 else 0 end),",
        "    times_rejected = times_rejected + (case when NEW.action = 'rejected' then 1 else 0 end),",
        "    avg_user_rating = (",
        "      select avg(rating)::double precision from venue_feedback",
        "      where venue_id = NEW.venue_id and rating is not null",
        "    ),",
        "    popularity_score = (",
        "      (times_accepted + 1.0) / (times_recommended + 2.0) * 100",
        "    ),",
        "    last_recommended = case when NEW.action in ('recommended','accepted') then now() else last_recommended end,",
        "    updated_at = now()",
        "  where id = NEW.venue_id;",
        "  return NEW;",
        "end;",
        "$$ language plpgsql;",
        "",
        "create trigger trg_venue_feedback_update",
        "  after insert on venue_feedback",
        "  for each row",
        "  execute function update_venue_popularity();",
        "",
        "-- ─── Seed all venues from Excel guides ────────────────────",
        "",
    ]

    # Insert in batches of 50
    batch_size = 50
    for i in range(0, len(venues), batch_size):
        batch = venues[i:i + batch_size]
        lines.append(f"insert into venues (id, name, slug, city, state, neighborhood, address, lat, lng, cuisine, cuisine_tags, vibe_tags, occasion_tags, price, price_level, vibe_notes, source_credit) values")

        values = []
        for v in batch:
            cuisine_arr = "ARRAY[" + ",".join(f"'{escape_sql(t)}'" for t in v["cuisine_tags"]) + "]::text[]" if v["cuisine_tags"] else "'{}'::text[]"
            vibe_arr = "ARRAY[" + ",".join(f"'{escape_sql(t)}'" for t in v["vibe_tags"]) + "]::text[]" if v["vibe_tags"] else "'{}'::text[]"
            occasion_arr = "ARRAY[" + ",".join(f"'{escape_sql(t)}'" for t in v["occasion_tags"]) + "]::text[]" if v["occasion_tags"] else "'{}'::text[]"

            values.append(
                f"  ('{escape_sql(v['id'])}', '{escape_sql(v['name'])}', '{escape_sql(v['slug'])}', "
                f"'{escape_sql(v['city'])}', '{escape_sql(v['state'])}', '{escape_sql(v['neighborhood'])}', "
                f"'{escape_sql(v['address'])}', {v['lat']}, {v['lng']}, '{escape_sql(v['cuisine'])}', "
                f"{cuisine_arr}, {vibe_arr}, {occasion_arr}, "
                f"'{escape_sql(v['price'])}', {v['price_level']}, '{escape_sql(v['vibe_notes'])}', "
                f"'{escape_sql(v['source_credit'])}')"
            )

        lines.append(",\n".join(values))
        lines.append("on conflict (id) do update set")
        lines.append("  occasion_tags = array_cat(venues.occasion_tags, excluded.occasion_tags),")
        lines.append("  vibe_tags = array_cat(venues.vibe_tags, excluded.vibe_tags),")
        lines.append("  updated_at = now();")
        lines.append("")

    lines.append("-- Done! Your AI agents now have local venue knowledge.")
    return "\n".join(lines)


def generate_typescript(venues: list[dict]) -> str:
    """Generate a TypeScript module the agents can import"""
    cities = sorted(set(v["city"] for v in venues))
    city_counts = {c: len([v for v in venues if v["city"] == c]) for c in cities}

    ts = f'''/**
 * Confetti Venue Knowledge Base
 * Auto-generated from Excel guides — DO NOT EDIT MANUALLY
 * Generated: {datetime.now().isoformat()}
 * Total: {len(venues)} venues across {len(cities)} cities
 *
 * This module provides the AI agents with local venue intelligence.
 * The agents use this as a FIRST source before hitting Google/Foursquare APIs.
 */

export interface VenueKnowledge {{
  id: string;
  name: string;
  slug: string;
  city: string;
  state: string;
  neighborhood: string;
  address: string;
  lat: number;
  lng: number;
  cuisine: string;
  cuisineTags: string[];
  vibeTags: string[];
  occasionTags: string[];
  price: string;
  priceLevel: number;
  vibeNotes: string;
  sourceCredit: string;
}}

/** Cities with curated local knowledge */
export const COVERED_CITIES = {json.dumps(cities)} as const;

/** City venue counts */
export const CITY_VENUE_COUNTS: Record<string, number> = {json.dumps(city_counts, indent=2)};

/**
 * Query the local venue knowledge base.
 * Used by venue-discovery.ts BEFORE hitting external APIs.
 */
export function queryLocalVenues(params: {{
  city: string;
  occasion?: string;
  vibes?: string[];
  priceLevel?: number;
  cuisine?: string;
  limit?: number;
}}): VenueKnowledge[] {{
  let results = VENUE_KNOWLEDGE.filter(v =>
    v.city.toLowerCase() === params.city.toLowerCase()
  );

  if (params.occasion) {{
    const occ = params.occasion.toLowerCase();
    const withOccasion = results.filter(v =>
      v.occasionTags.some(t => t.includes(occ) || occ.includes(t))
    );
    if (withOccasion.length > 0) results = withOccasion;
  }}

  if (params.vibes?.length) {{
    const vibeSet = new Set(params.vibes.map(v => v.toLowerCase()));
    results.sort((a, b) => {{
      const aMatch = a.vibeTags.filter(t => vibeSet.has(t)).length;
      const bMatch = b.vibeTags.filter(t => vibeSet.has(t)).length;
      return bMatch - aMatch;
    }});
  }}

  if (params.priceLevel) {{
    results = results.filter(v => v.priceLevel <= params.priceLevel!);
  }}

  if (params.cuisine) {{
    const c = params.cuisine.toLowerCase();
    const withCuisine = results.filter(v =>
      v.cuisine.toLowerCase().includes(c) ||
      v.cuisineTags.some(t => t.toLowerCase().includes(c))
    );
    if (withCuisine.length > 0) results = withCuisine;
  }}

  return results.slice(0, params.limit ?? 15);
}}

/**
 * Check if we have local knowledge for a city.
 * If true, agents should prefer local data over API-only results.
 */
export function hasLocalKnowledge(city: string): boolean {{
  return VENUE_KNOWLEDGE.some(v => v.city.toLowerCase() === city.toLowerCase());
}}

/** Full venue knowledge array */
export const VENUE_KNOWLEDGE: VenueKnowledge[] = {json.dumps([
    {
        "id": v["id"],
        "name": v["name"],
        "slug": v["slug"],
        "city": v["city"],
        "state": v["state"],
        "neighborhood": v["neighborhood"],
        "address": v["address"],
        "lat": v["lat"],
        "lng": v["lng"],
        "cuisine": v["cuisine"],
        "cuisineTags": v["cuisine_tags"],
        "vibeTags": v["vibe_tags"],
        "occasionTags": v["occasion_tags"],
        "price": v["price"],
        "priceLevel": v["price_level"],
        "vibeNotes": v["vibe_notes"],
        "sourceCredit": v["source_credit"],
    }
    for v in venues
], indent=2)};
'''
    return ts


def main():
    all_venues = []
    files = sorted(glob.glob(GUIDE_PATTERN))

    if not files:
        print("❌ No guide files found!")
        return

    print(f"🎊 Confetti Venue Seeder")
    print(f"{'='*50}")
    print(f"Found {len(files)} guide files\n")

    for f in files:
        venues = parse_workbook(f)
        print(f"  📄 {Path(f).name}: {len(venues)} venues")
        all_venues.extend(venues)

    # Deduplicate (merge occasion tags for same venue)
    deduped = deduplicate(all_venues)
    print(f"\n📊 Total: {len(all_venues)} raw → {len(deduped)} unique venues")
    print(f"   (Merged {len(all_venues) - len(deduped)} duplicates, kept all occasion tags)")

    # Generate SQL migration
    migration_dir = ROOT / "supabase" / "migrations"
    migration_dir.mkdir(parents=True, exist_ok=True)
    migration_path = migration_dir / "20260524000000_seed_venue_knowledge.sql"
    migration_path.write_text(generate_migration(deduped))
    print(f"\n✅ SQL migration: {migration_path.relative_to(ROOT)}")

    # Generate TypeScript module
    ts_path = ROOT / "src" / "lib" / "agents" / "venue-knowledge.ts"
    ts_path.write_text(generate_typescript(deduped))
    print(f"✅ TypeScript:     {ts_path.relative_to(ROOT)}")

    # Summary
    from collections import Counter
    city_counts = Counter(v["city"] for v in deduped)
    occasion_counts = Counter(t for v in deduped for t in v["occasion_tags"])

    print(f"\n📍 Cities ({len(city_counts)}):")
    for city, count in city_counts.most_common():
        print(f"   {city}: {count}")

    print(f"\n🎯 Occasions ({len(occasion_counts)}):")
    for occ, count in occasion_counts.most_common(10):
        print(f"   {occ}: {count}")

    print(f"""
{'='*50}
🚀 NEXT STEPS:
{'='*50}

1. Push to Lovable (the TypeScript file is already in your src/):
   → src/lib/agents/venue-knowledge.ts

2. Apply the Supabase migration:
   → supabase/migrations/20260524000000_seed_venue_knowledge.sql

3. Wire into venue-discovery.ts:
   Import queryLocalVenues() and call it BEFORE the Google/Foursquare APIs.
   When local knowledge exists for a city, blend it with API results.

4. The learning loop:
   → venue_feedback table tracks accepted/rejected/rated
   → popularity_score auto-updates via trigger
   → Agents prioritize high-popularity venues over time
""")


if __name__ == "__main__":
    main()
