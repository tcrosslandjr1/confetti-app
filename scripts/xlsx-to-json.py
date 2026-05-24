#!/usr/bin/env python3
"""
Convert all Confetti Guide .xlsx files into a single venues.json
that can be imported into the app or seeded to Supabase.

Usage:
  python3 scripts/xlsx-to-json.py

Output:
  scripts/venues.json  — array of venue objects ready for DB insert
"""

import os, json, re, glob
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parent.parent
GUIDE_PATTERN = str(ROOT / "*_Guide*.xlsx")

def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")

def price_to_level(price: str) -> int:
    if not price:
        return 2
    count = price.count("$")
    return min(max(count, 1), 4)

def extract_city_from_filename(filename: str) -> str:
    """Pull city/region name from filename like 'DC_Confetti_Guide.xlsx'"""
    base = Path(filename).stem
    # Remove common suffixes
    for suffix in ["_Confetti_Guide", "_Guide", "_Spots", "_v2"]:
        base = base.replace(suffix, "")
    # Convert underscores to spaces
    return base.replace("_", " ").strip()

def parse_workbook(filepath: str) -> list[dict]:
    venues = []
    city = extract_city_from_filename(filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        # Skip overview/summary sheets
        if sheet_name.lower() in ("overview", "summary", "notes", "sources"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # First row is header
        header = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Map column indices
        col_map = {}
        for i, h in enumerate(header):
            if "name" in h:
                col_map["name"] = i
            elif "neighbor" in h:
                col_map["neighborhood"] = i
            elif "address" in h:
                col_map["address"] = i
            elif "type" in h or "cuisine" in h:
                col_map["cuisine"] = i
            elif "price" in h:
                col_map["price"] = i
            elif "vibe" in h or "note" in h:
                col_map["vibe"] = i
            elif "source" in h:
                col_map["source"] = i

        if "name" not in col_map:
            continue

        for row in rows[1:]:
            name = row[col_map["name"]] if col_map.get("name") is not None and row[col_map["name"]] else None
            if not name:
                continue

            name = str(name).strip()
            neighborhood = str(row[col_map.get("neighborhood", 0)] or "").strip()
            address = str(row[col_map.get("address", 0)] or "").strip()
            cuisine = str(row[col_map.get("cuisine", 0)] or "").strip()
            price = str(row[col_map.get("price", 0)] or "$$").strip()
            vibe = str(row[col_map.get("vibe", 0)] or "").strip()
            source = str(row[col_map.get("source", 0)] or "").strip()

            # Build tags from cuisine + vibe keywords
            tags = [t.strip().lower() for t in cuisine.split("/") if t.strip()]
            tags.append(slugify(sheet_name))

            venue = {
                "name": name,
                "slug": slugify(name),
                "city": city,
                "neighborhood": neighborhood,
                "address": address,
                "cuisine": cuisine,
                "price": price,
                "price_level": price_to_level(price),
                "vibe_notes": vibe,
                "source": source,
                "category": sheet_name.strip(),
                "tags": tags,
            }
            venues.append(venue)

    wb.close()
    return venues


def main():
    all_venues = []
    files = sorted(glob.glob(GUIDE_PATTERN))

    if not files:
        print("No guide files found! Check that .xlsx files are in project root.")
        return

    print(f"Found {len(files)} guide files:\n")
    for f in files:
        venues = parse_workbook(f)
        print(f"  {Path(f).name}: {len(venues)} venues")
        all_venues.append({"file": Path(f).name, "city": extract_city_from_filename(f), "count": len(venues)})
        all_venues_flat = [] if not hasattr(main, '_venues') else main._venues

    # Actually collect all venues
    all_venues_flat = []
    for f in files:
        all_venues_flat.extend(parse_workbook(f))

    # Deduplicate by name+city
    seen = set()
    deduped = []
    for v in all_venues_flat:
        key = (v["name"].lower(), v["city"].lower())
        if key not in seen:
            seen.add(key)
            deduped.append(v)

    output_path = ROOT / "scripts" / "venues.json"
    output_path.parent.mkdir(exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(deduped, f, indent=2)

    print(f"\n✅ Exported {len(deduped)} unique venues to {output_path}")
    print(f"   (Removed {len(all_venues_flat) - len(deduped)} duplicates)")

    # Also print city breakdown
    from collections import Counter
    city_counts = Counter(v["city"] for v in deduped)
    print(f"\n📍 Cities: {len(city_counts)}")
    for city, count in city_counts.most_common():
        print(f"   {city}: {count}")


if __name__ == "__main__":
    main()
